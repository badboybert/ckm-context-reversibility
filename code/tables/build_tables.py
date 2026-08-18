#!/usr/bin/env python
"""build_tables.py — Paper B manuscript tables workbook (Table 1 determinant + Table 2 cohort + supplementary sheets).
All values are analysis outputs read from results/ (data tables, not a financial model -> values are the data).
Output: manuscript/tables/Paper_B_Tables.xlsx"""
import os, math, pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
HERE=os.path.dirname(os.path.abspath(__file__)); SIG=os.path.dirname(os.path.dirname(HERE))
RES=os.path.join(SIG,'results'); RD=lambda p: pd.read_csv(os.path.join(RES,p),sep='\t')

HDR_FILL=PatternFill('solid',fgColor='1F3864'); HDR_FONT=Font(name='Arial',bold=True,color='FFFFFF',size=10)
TTL_FONT=Font(name='Arial',bold=True,size=12,color='1F3864'); NOTE_FONT=Font(name='Arial',italic=True,size=8,color='595959')
BODY=Font(name='Arial',size=9); THIN=Side(style='thin',color='D9D9D9'); BORDER=Border(bottom=THIN)
CENTER=Alignment(horizontal='center',vertical='center'); LEFT=Alignment(horizontal='left',vertical='center',wrap_text=True)

def write_sheet(wb, name, title, note, df, widths, center_cols=()):
    ws=wb.create_sheet(name)
    ws['A1']=title; ws['A1'].font=TTL_FONT
    ws['A2']=note;  ws['A2'].font=NOTE_FONT; ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=max(2,len(df.columns)))
    hr=4
    for j,col in enumerate(df.columns,1):
        c=ws.cell(hr,j,col); c.fill=HDR_FILL; c.font=HDR_FONT; c.alignment=CENTER; c.border=BORDER
    for i,(_,row) in enumerate(df.iterrows(),hr+1):
        for j,col in enumerate(df.columns,1):
            v=row[col]; c=ws.cell(i,j,'' if pd.isna(v) else v); c.font=BODY; c.border=BORDER
            c.alignment=CENTER if col in center_cols else LEFT
    for j,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(j)].width=w
    ws.freeze_panes=ws.cell(hr+1,1)
    ws.auto_filter.ref=f"A{hr}:{get_column_letter(len(df.columns))}{hr+len(df)}"
    return ws

wb=Workbook(); wb.remove(wb.active)

# ---------- Table 2: cohort & design inventory ----------
m=RD('panel_manifest_full.tsv').copy()
m['n_pairs']=m['n_pairs'].apply(lambda x:'—' if x is not None and x<0 else x)   # sig-only corroboration panels: n unknown
m['contributes_to_score']=m['contributes_to_score'].map({True:'Y',False:'',1:'Y',0:''}).fillna('')
m['persistence_eligible']=m['persistence_eligible'].map({True:'Y',False:'',1:'Y',0:''}).fillna('')
# Determinant-eligibility routing (which panels feed which analysis; the k=9/k=3
# determinant contexts were previously enumerated only in the Methods). Derived, not hardcoded.
_dp=RD('determinant_per_panel.tsv')
_det={p: lay for p, lay in _dp.drop_duplicates('panel').set_index('panel')['layer'].items()}
m['determinant']=m['panel'].map(lambda p: {'transcriptome':'Y (k=9)','proteome':'Y (k=3)'}.get(_det.get(p),''))
assert (m['determinant']=='Y (k=9)').sum()==9 and (m['determinant']=='Y (k=3)').sum()==3, \
    f"determinant routing mismatch: {m['determinant'].value_counts().to_dict()}"
# Reader-facing label maps (journal-facing strings; the internal snake_case stays in the manifest).
_LAYER_R={'proteome':'Proteome','transcriptome':'Transcriptome','methylation':'Methylome','metabolome':'Metabolome'}
_PLAT_R={'RNAseq':'RNA-seq','array':'microarray','microarray':'microarray','HuGene':'microarray (HuGene)',
         '450K':'450K array','EPIC':'EPIC array','MS':'mass spectrometry','Olink':'Olink','SomaScan':'SomaScan',
         'Metabolon':'Metabolon','Biocrates':'Biocrates','LC-HRMS':'LC-HRMS'}
_TIS_R={'SAT_adipose':'subcutaneous adipose','skeletal_muscle':'skeletal muscle','whole_blood':'whole blood',
        'blood_neutrophil':'neutrophils','blood':'blood (bulk)','PBMC':'PBMC','plasma':'plasma','serum':'serum','liver':'liver'}
_STAT_R={'ANALYZED':'Analysed','corroboration-only':'Corroboration only','RETAINED-memory-comparator':'Retained comparator',
         'EXCLUDED':'Excluded'}
def _mapcol(s,d): return s.map(lambda v: d.get(v,v))

# ---------- Table 2 (main, COMPACT): one reader-facing row per molecular layer ----------
# The 17-column registry is a provenance object, not a reader-facing main table (round-5 review);
# it moves to the 'Table 2 - full registry' sheet below. Counts here are DERIVED from the manifest and
# asserted, never hand-typed (the 51.47->52 lesson): a per-layer summary that silently drifts from the
# registry would be worse than none.
_an=m[m['status'].astype(str).str.upper()!='EXCLUDED'].copy()
_layorder=['proteome','transcriptome','methylation','metabolome']
_families={  # intervention-family reader summary per layer (qualitative grouping of the manifest's intervention column)
 'proteome':'diet; bariatric surgery; GLP-1 receptor agonist; SGLT2 inhibitor; metformin',
 'transcriptome':'caloric restriction & diet; bariatric surgery; exercise; GLP-1 receptor agonist; metformin',
 'methylation':'diet; bariatric surgery; exercise',
 'metabolome':'bariatric surgery'}
_design={'proteome':'within-person & randomized-trial contrasts','transcriptome':'mostly within-person (some trial contrasts)',
 'methylation':'within-person','metabolome':'published within-person effects'}
_roles={'proteome':'score; restoration; drug comparison; durability','transcriptome':'primary determinant test; context; durability',
 'methylation':'replication; reversible/persistent pole asymmetry','metabolome':'score; restoration'}
_limit={'proteome':'platform & estimand heterogeneity','transcriptome':'panel heterogeneity; bulk-tissue composition',
 'methylation':'q-value-only scoring; age-related drift','metabolome':'few contexts; descriptive'}
_rows=[]
for lay in _layorder:
    s=_an[_an['layer']==lay]
    npan=int((s['status'].astype(str).str.upper()=='ANALYZED').sum())
    comps='; '.join(sorted({_TIS_R.get(t,t) for t in s['tissue'].dropna().unique()}))
    _rows.append({'Layer':_LAYER_R[lay],'Panels (n)':npan,'Compartments':comps,
                  'Interventions':_families[lay],'Design / estimand':_design[lay],
                  'Primary roles':_roles[lay],'Key limitation':_limit[lay]})
t2c=pd.DataFrame(_rows)
assert t2c['Panels (n)'].sum()==40, f"compact Table 2 panel counts sum to {t2c['Panels (n)'].sum()}, not the 40 analysed"
write_sheet(wb,'Table 2','Table 2 | Cohort and study-design summary',
  'Reader-facing overview of the 40 analysed human intervention panels, one row per molecular layer; the full per-panel registry (accessions, platforms, '
  'sample sizes, analysis-routing flags and provenance notes) is the "Table 2 - full registry" sheet. Panel counts are the analysed panels per layer and sum to 40. '
  'Compartments and interventions are grouped to reader-facing families. Source: supplementary_tables/panel_manifest_full.tsv.',
  t2c,[14,10,30,40,32,34,30],center_cols={'Panels (n)'})

# ---------- Table 2 - full registry (supplementary provenance): the 17-column per-panel table ----------
ed1=m[['panel','layer','platform','intervention','tissue','n_pairs','n_marks','sig_marks','sig_basis',
       'bmi_axis','base_sd','role','status','contributes_to_score','persistence_eligible','determinant','note']].copy()
ed1['layer']=_mapcol(ed1['layer'],_LAYER_R); ed1['platform']=_mapcol(ed1['platform'],_PLAT_R)
ed1['tissue']=_mapcol(ed1['tissue'],_TIS_R); ed1['status']=_mapcol(ed1['status'],_STAT_R)
ed1.columns=['Panel / accession','Layer','Platform','Intervention','Tissue','n (pairs)','n marks','Sig. marks',
             'Sig. basis','BMI axis','Baseline SD','Role','Status','Scores','Persist.','Determinant','Note']
ed1=ed1.sort_values(['Layer','Tissue','Panel / accession'])
write_sheet(wb,'Table 2 - full registry','Table 2 (full registry) | Per-panel cohort and study-design inventory',
  'Every human intervention panel analysed (40 analysed; 4 layers; 26 contribute to the reversibility score, 8 persistence-eligible). '
  'The Role, Status, Scores, Persist., BMI axis and Determinant columns route each panel to the analyses it enters: Scores = contributes to the reversibility score; '
  'Persist. = persistence-eligible; BMI axis = restoration-eligible (and the axis source); Determinant = enters the determinant meta-analysis, marked Y (k=9) for the '
  'powered transcriptome contexts and Y (k=3) for the inconclusive plasma-proteome contexts. '
  'Proteome baseline SD = population-variance proxy (UKB-PPP/deCODE), caveated. GSE199063 = Affymetrix Clariom-D (AceView novel transcripts excluded from gene-level analyses). '
  'Source: supplementary_tables/panel_manifest_full.tsv; supplementary_tables/determinant_per_panel.tsv.',
  ed1,[26,13,16,15,18,9,9,9,12,20,16,17,14,8,9,12,40],
  center_cols={'Layer','Platform','n (pairs)','n marks','Sig. marks','Scores','Persist.','Determinant'})

# ---------- Table 2: determinant model ----------
NAME={'loeuf':'Genetic constraint (LOEUF)','n_drug_log':'Druggability (log)','n_gwas_log':'GWAS burden (log)',
 'tau':'Tissue-specificity (τ)','arch_nsig':'# cis-eQTL signals','arch_str':'cis-eQTL strength (F)','has_arch':'Has cis-QTL',
 'causal_nonEGFR':'Genetic causal status','is_enzyme':'Enzyme','is_membrane':'Membrane','is_secreted':'Secreted',
 'loeuf_miss':'LOEUF missing (indicator)','tau_miss':'τ missing (indicator)'}
dm=RD('determinant_meta.tsv').copy()
dm['Feature']=dm['feature'].map(NAME)
dm['Pooled β (SD)']=dm['pooled_beta'].round(4)
dm['95% CI']=dm.apply(lambda r:f"({r.pooled_beta-r.ci_halfwidth:+.4f}, {r.pooled_beta+r.ci_halfwidth:+.4f})",axis=1)
dm['MDE (SD)']=dm['mde'].round(4)
dm['I² (%)']=dm['I2'].round(1)
dm['τ²']=dm['tau2'].apply(lambda x:'0' if x==0 else f"{x:.2e}")
dm['Cochran Q (df)']=dm.apply(lambda r:f"{r.Q:.2f} ({int(r.df)})",axis=1)
dm['n genes (median)']=dm['n_genes_median'].astype(int)
def _predicts(r):
    # A "no" verdict is only defensible where the interval can actually EXCLUDE a meaningful effect.
    # The proteome causal-status row previously read "no (|β| < 0.03)" on a CI of (-0.069, +0.089)
    # and an MDE of 0.079 — i.e. it asserted a null the data cannot support, which is the same
    # underpowered-null-as-equivalence error the paper corrects elsewhere. The verdict is therefore
    # driven by CI containment within the ±0.05 SESOI, not by |β| alone.
    # The rare-binary carve-out is TRANSCRIPTOME-ONLY, and the layer condition is load-bearing:
    # causal status is carried by 9 of 18,615 transcripts (0.05%, SD_x~0.022 -> the ~50x compression S14
    # documents) but by 123 of 1,463 proteins (8.4%, SD_x~0.28 -> no meaningful compression). The proteome
    # row must therefore fall through to the CI test and read "inconclusive", like every other proteome
    # coefficient: its problem is the underpowered k=3 arm (MDE 0.079), not standardization.
    if r['feature']=='causal_nonEGFR' and r['layer']=='transcriptome':
        return 'underpowered rare-binary (see S14)'
    ci_lo, ci_hi = r.pooled_beta - r.ci_halfwidth, r.pooled_beta + r.ci_halfwidth
    if r['inconclusive'] or ci_lo < -0.05 or ci_hi > 0.05:
        # "underpowered" was applied to every inconclusive row, including tissue-specificity, whose
        # MDE (0.035) sits INSIDE the SESOI -- it is not underpowered, its interval is widened by
        # I2 = 96.9%. Name which of the three reasons actually applies, from this row's own numbers.
        if r.mde > 0.05:
            return 'inconclusive (underpowered: MDE exceeds SESOI)'
        if r.I2 >= 50:
            return 'inconclusive (heterogeneous; CI breaches SESOI)'
        return 'inconclusive (CI breaches SESOI)'
    return 'no (95% CI within ±0.05 SESOI)' if abs(r.pooled_beta)<0.03 else 'small effect'
dm['Practical-equivalence verdict']=dm.apply(_predicts,axis=1)
dm['Layer (k panels)']=dm['layer']+' (k='+dm['k'].astype(str)+')'
# Transcriptome (the powered, primary test) first, then the inconclusive proteome block, so the
# main table reads primary-then-supporting instead of alphabetically by layer.
_LAYER_ORDER={'transcriptome':0,'proteome':1}
dm['_lo']=dm['layer'].map(_LAYER_ORDER)
assert dm['_lo'].notna().all(), f"unexpected layer(s): {sorted(set(dm.loc[dm['_lo'].isna(),'layer']))}"
dm=dm.sort_values(['_lo','Feature'])
# tau2 and Cochran Q move to their own sheet: 10 columns do not fit a portrait Word table, and I2
# is the interpretable heterogeneity summary for a reader. Nothing is dropped, only relocated.
t2=dm[['Feature','Layer (k panels)','Pooled β (SD)','95% CI','MDE (SD)','I² (%)','n genes (median)','Practical-equivalence verdict']]
write_sheet(wb,'Table 1','Table 1 | Portable mark-intrinsic features do not predict practically meaningful transcriptome reversibility (determinant meta-analysis)',
  "Standardized determinant coefficients (HC3 OLS per panel, measurability-residualized outcome; Hartung-Knapp random-effects meta across panels). "
  "Transcriptome (k=9 contexts / 4 tissues) is the powered test: across the continuous mark-intrinsic features every |pooled β| < 0.029 SD against median MDE 0.013 SD, none reaching the ±0.05 SD SESOI. Verdicts are assigned by containment of the 95% CI within the ±0.05 SD SESOI, not by |β| alone: tissue-specificity (τ) has |β| = 0.026 but its interval reaches +0.061, so it reads inconclusive rather than equivalent, matching Supplementary Table 9. "
  "Genetic causal status is a rare 0/1 feature (9 transcripts); its standardized coefficient is compressed by the small predictor variance and is not the primary interpretable contrast for this rare binary predictor and is not evaluated against the continuous-feature equivalence bound (interpretable raw contrast −0.18 SD, 95% CI −0.37 to +0.01, underpowered rather than equivalent; Supplementary Table 14). "
  "Proteome (k=3) is reported as inconclusive (wide CIs) and is shown as a separate block below the transcriptome rows. Between-panel dispersion is propagated into each pooled estimate through Hartung-Knapp random-effects confidence intervals; per-feature τ² and Cochran's Q are in the \u2018Table 1 \u2014 heterogeneity\u2019 sheet. Source: supplementary_tables/determinant_meta.tsv.",
  t2,[30,20,13,21,11,9,14,34],
  center_cols={'Pooled β (SD)','MDE (SD)','Layer (k panels)','I² (%)','n genes (median)'})

# ---------- Table 1 heterogeneity detail (moved out of the main table) ----------
t1h=dm[['Feature','Layer (k panels)','I² (%)','τ²',"Cochran Q (df)",'MDE (SD)']]
write_sheet(wb,'Table 1 - heterogeneity','Table 1 (detail) | Between-panel heterogeneity for each determinant coefficient',
  "Heterogeneity statistics for the same feature x layer rows as main Table 1, moved here to keep the main table readable at portrait width. "
  "I2 is the proportion of between-panel variance not attributable to sampling error; tau-squared is its absolute scale; Cochran's Q is tested on k-1 degrees of freedom. "
  "Heterogeneity was near-absent (I2 = 0%) for genetic causal status, the number of cis-eQTL signals, cis-instrument strength and the presence of a cis-QTL, and substantial for druggability (I2 = 57.8%) and tissue-specificity (I2 = 96.9%). "
  "Dispersion is already propagated into every pooled interval in main Table 1 through the Hartung-Knapp random-effects estimator, so these columns explain the width of those intervals rather than adding a separate test. "
  "Source: supplementary_tables/determinant_meta.tsv.",
  t1h,[30,20,9,11,13,11],
  center_cols={'Layer (k panels)','I² (%)','τ²',"Cochran Q (df)",'MDE (SD)'})

# ---------- Supp: universal reversible core ----------
core=RD('universal_reversible_core.tsv').copy()
core.columns=['Gene','Tissue pairs shared (of 6)']
write_sheet(wb,'S1 - universal core','Supplementary Table 1 | Universal cross-tissue reversible core (100 genes)',
  'Genes in the strongly-reversible consensus of >=2 tissues (union of cross-tissue reversible-consensus intersections). '
  'Inflammation/innate-immune + metabolic-remodelling programme. Source: supplementary_tables/universal_reversible_core.tsv.',
  core,[16,26],center_cols={'Tissue pairs shared (of 6)'})

# ---------- Supp: drug-class correlations ----------
dc=RD('drug_class_corr.tsv').copy()
dc=dc[pd.notna(dc['rho'])][['a','b','n','rho','p','ca','cb']].copy()
# Carry the engine's OWN powered flag onto the sheet. The note used to say "powered pairs only, n>=30"
# while the sheet filtered nothing: dapagliflozin x semaglutide (n=34, rho=+0.42) therefore sat at the
# TOP of a rho-sorted sheet whose note said empagliflozin is the only adequately powered SGLT2-inhibitor
# panel. The engine flags that pair powered_both=False and Figure 5 excludes it. Label, don't hide.
_rec=RD('drug_class_reconciliation.tsv')[['pair_a','pair_b','powered_both']]
_pw={}
for r in _rec.itertuples(index=False):
    _pw[frozenset((r.pair_a, r.pair_b))]=bool(r.powered_both)
dc['powered']=[ ('Yes' if _pw.get(frozenset((a,b)), False) else 'No') for a,b in zip(dc['a'],dc['b']) ]
dc=dc[['a','b','n','rho','p','ca','cb','powered']]
dc.columns=['Intervention A','Intervention B','n shared','Spearman ρ','P','Class A','Class B','Powered (both panels)']
dc['Spearman ρ']=dc['Spearman ρ'].round(3); dc['P']=dc['P'].apply(lambda x:f"{x:.1e}")
dc=dc.sort_values('Spearman ρ',ascending=False)
write_sheet(wb,'S2 - drug class','Supplementary Table 2 | Cross-intervention plasma-proteome reversal correlations',
  'Pairwise Spearman of reversal effects across every intervention pair with an estimable correlation. All pairs are listed; the "Powered (both panels)" column carries the analysis\'s own '
  'power flag, and only powered pairs enter Figure 5 and the inferences below. In the available powered plasma-proteome panels, semaglutide\'s reversal signature tracks the magnitude of weight '
  'loss — bariatric surgery (+0.32, 95% CI +0.27 to +0.37, n=1,253) and diet (+0.31, +0.28 to +0.33, n=4,113) — but not empagliflozin (+0.05, -0.01 to +0.11, n=1,115, ns); those differences are '
  'powered as contrasts (Fisher z = 6.8 and 7.8). This is not a claim that drug class is absent: the within-pharmacotherapy pairs are too few (k=3 informative) to support an equivalence bound, and this '
  'sheet contains a counter-example that is unpowered but should be read rather than hidden — dapagliflozin x semaglutide (rho=+0.42, P=0.013) is the largest rho in the table, but rests on only '
  '34 jointly measured proteins, is flagged not-powered, and is excluded from Figure 5; dapagliflozin is a corroboration-only panel (1-34 shared proteins across its pairs). Empagliflozin is the '
  'only adequately powered SGLT2-inhibitor panel, so no SGLT2-class-level conclusion is drawn from either. '
  'Cross-platform >= same-platform => not a platform artifact. Source: supplementary_tables/drug_class_corr.tsv; power flag from supplementary_tables/drug_class_reconciliation.tsv.',
  dc,[16,16,10,12,12,12,12,15],center_cols={'n shared','Spearman ρ','P','Class A','Class B','Powered (both panels)'})

# ---------- Supp: per-context reversibility atlas ----------
at=RD('context_signatures.tsv').copy()
at=at[['layer','context','tissue','intervention','rank','mark','rev_beta']]
at['rev_beta']=at['rev_beta'].round(3)
at.columns=['Layer','Context','Tissue','Intervention','Rank','Gene/mark','Reversal β']
write_sheet(wb,'S3 - context atlas','Supplementary Table 3 | Per-context reversibility atlas (top-25 reversers × 10 contexts)',
  'The strongest-reversing marks in each tissue×intervention context (rank 1–25 by |reversal β|). Source: supplementary_tables/context_signatures.tsv.',
  at,[14,26,12,20,8,14,12],center_cols={'Rank','Reversal β','Layer'})

# ---------- Supp: per-intervention systemic consensus ----------
iv2=RD('intervention_systemic_consensus.tsv').copy()
iv2.columns=['Intervention','n panels','n tissues','Cross-tissue consensus genes','Top genes']
write_sheet(wb,'S4 - intervention consensus','Supplementary Table 4 | Per-intervention cross-tissue consensus',
  'Genes reversed by an intervention across >=half its tissue panels. Surgery/diet write systemic (cross-tissue) signatures; exercise/drug do not. Source: supplementary_tables/intervention_systemic_consensus.tsv + consensus_signatures.txt.',
  iv2,[16,10,10,28,60],center_cols={'n panels','n tissues','Cross-tissue consensus genes'})

# ---------- Supp: context structure (tissue >> intervention) ----------
cs=RD('cross_context_summary.txt') if False else None  # txt, not tsv; use the F6 source CSVs
import pandas as _pd
fa=_pd.read_csv(os.path.join(HERE,'..','figures','source_data','F6a_context_corr.csv'))
fb=_pd.read_csv(os.path.join(HERE,'..','figures','source_data','F6b_concordance.csv'))
mat=fa.pivot(index='ctx_a',columns='ctx_b',values='rho').round(3).reset_index().rename(columns={'ctx_a':'Context'})
write_sheet(wb,'S5 - context structure','Supplementary Table 5 | Reversibility is tissue-organized (context correlation + concordance)',
  'Top: genome-wide rev_beta Spearman between the 4 powered transcriptome contexts, computed on the genes measured in BOTH contexts of each pair (within-adipose +0.52, cross-tissue ~0). '
  'The lower block reports, per pair, that shared MEASURED universe (the denominator of the genome-wide rho) alongside the shared SIGNIFICANT count (the denominator of the directional concordance) — '
  'the two are different numbers and are not interchangeable. Sources: supplementary_tables/cross_context_summary.txt; manuscript/figures/source_data/F6{a,b}_*.csv.',
  mat,[16,12,12,12,12],center_cols=set(mat.columns)-{'Context'})
fb['concordance']=fb['concordance'].round(3)
fb=fb[['pair','n_measured_both','n_shared','concordance','kind']]
fb.columns=['Context pair','n genes measured in both','n shared sig','Directional concordance','Kind']
ws_cs=wb['S5 - context structure']; r0=4+len(mat)+2
ws_cs.cell(r0,1,'Directional concordance (same- vs cross-tissue):').font=NOTE_FONT
for j,col in enumerate(fb.columns,1):
    c=ws_cs.cell(r0+1,j,col); c.fill=HDR_FILL; c.font=HDR_FONT; c.alignment=CENTER
for i,(_,row) in enumerate(fb.iterrows(),r0+2):
    for j,col in enumerate(fb.columns,1):
        c=ws_cs.cell(i,j,row[col]); c.font=BODY; c.alignment=CENTER if col!='Context pair' else LEFT

# ---------- Supp: predictive null ----------
pn=RD('predictive_null.tsv').copy(); ab=RD('predictive_null_ablation.tsv').copy()
pn=pn.merge(ab,on='context',how='left')
for c in ['within_spearman','loco_spearman','within_auc','loco_auc','loco_auc_full','loco_auc_intrinsic_only','loco_auc_measurability_only','pos_rate']:
    if c in pn: pn[c]=pn[c].round(3)
pn=pn[['context','n','within_spearman','loco_spearman','within_auc','loco_auc','loco_auc_intrinsic_only','loco_auc_measurability_only']]
pn.columns=['Context','n genes','Within ρ','LOCO ρ','Within AUC','LOCO AUC','LOCO AUC (intrinsic only)','LOCO AUC (measurability only)']
write_sheet(wb,'S6 - predictive null','Supplementary Table 6 | Predictive null (within-context vs leave-one-context-out)',
  'Gradient-boosted reversibility prediction: learns within a context (within ρ +0.13, perm→0) but does not transfer (LOCO ρ +0.002). The only cross-context signal is detectability (measurability-only ablation). Source: supplementary_tables/predictive_null{,_ablation}.tsv.',
  pn,[14,9,11,11,11,11,22,24],center_cols={'n genes','Within ρ','LOCO ρ','Within AUC','LOCO AUC','LOCO AUC (intrinsic only)','LOCO AUC (measurability only)'})

# ---------- Supp: durable reversers (canonical) ----------
import numpy as _np, re as _re2
dur=RD('durability_GSE199063.tsv').dropna(subset=['rev2','rev5'])
NC2=_re2.compile(r'-AS\d?|^LINC|orf\d|^RNU|^SNOR|^MIR\d|^LOC|^AC\d|^AL\d|^RP[0-9LS]|^IG[HKL][VJC]')
g2=dur['gene'].astype(str); dur=dur[g2.str.match(r'^[A-Z][A-Z0-9-]*$') & ~g2.str.contains(NC2,na=False)]
sigd=dur[(dur.p2<0.05)&(dur.rev2.abs()>0.2)].copy(); sigd['ratio']=sigd.rev5/sigd.rev2
sigd['Class']=_np.where(sigd.ratio>0.7,'durable',_np.where(sigd.ratio<0.3,'rebound','partial'))
durtab=sigd.reindex(sigd.rev2.abs().sort_values(ascending=False).index).head(60)[['gene','rev2','rev5','ratio','Class']].copy()
durtab[['rev2','rev5','ratio']]=durtab[['rev2','rev5','ratio']].round(3)
durtab.columns=['Gene','2-yr reversal','5-yr reversal','5yr/2yr ratio','Class']
write_sheet(wb,'S7 - durable reversers','Supplementary Table 7 | Adipose durable vs rebound reversers (RYGB, 2yr→5yr; canonical genes)',
  'Top-60 significant 2-yr reversers (canonical HGNC genes; GSE199063 Affymetrix, AceView novel transcripts excluded). durable = 5yr/2yr ratio>0.7; rebound <0.3. Source: supplementary_tables/durability_GSE199063.tsv.',
  durtab,[14,14,14,14,10],center_cols={'2-yr reversal','5-yr reversal','5yr/2yr ratio','Class'})

# =====================================================================================
# NEW SUPPLEMENTARY SHEETS S8-S13 (Workstream B robustness / causal-status)
# =====================================================================================

def _sci(x):
    try: v=float(x)
    except (TypeError,ValueError): return x
    return f"{v:.2e}"

# A P value is never exactly zero; a printed "0.00e+00" is an artifact of how it was obtained, and the
# honest replacement differs by METHOD. Keep these two apart — see the S16 note below.
RESTORATION_B = 10000   # restoration_uncertainty.py L69 (B); perm P is a bare mean, so it is >= 1/B
def _sci_analytic(x):
    """Closed-form P (e.g. exact binomial): a 0 is double underflow."""
    try: v=float(x)
    except (TypeError,ValueError): return x
    return "<1e-300" if v==0 else f"{v:.2e}"
def _sci_perm(x, B=RESTORATION_B):
    """Empirical permutation P over B draws: bounded below by 1/B, NOT by machine epsilon."""
    try: v=float(x)
    except (TypeError,ValueError): return x
    return f"<{1.0/B:.0e}" if v==0 else f"{v:.2e}"
def _yn(x): return 'Yes' if bool(x) else 'No'

# ---------- S8: genetic causal-status screen (B2; 153 rows) ----------
cs8=RD('causal_status_supp_table.tsv').copy()
cs8['MR β']=cs8['MR_beta'].apply(lambda v:f"{v:+.4f}")
cs8['MR P']=cs8['MR_p'].apply(_sci); cs8['MR q']=cs8['q'].apply(_sci)
cs8['F-stat']=cs8['F'].round(1)
s8=cs8[['gene','layer','outcome','instrument_rsid','n_instr','F-stat','MR β','MR P','MR q',
        'coloc_PP_H4','PP_H4_over_H3','tier','n_nonEGFR_causal_outcomes','included_in_determinant',
        'determinant_panels_measured','in_PaperA_S2']].copy()
s8.columns=['Gene','Layer','Outcome','Instrument (cis)','n instr','F-stat','MR β','MR P','MR q',
            'Coloc PP.H4','PP.H4/PP.H3','Tier / MR gate','# non-eGFR causal outcomes','In determinant',
            'Panels measured','Coloc-staged in the companion cis-MR screen (ref. 17)']
write_sheet(wb,'S8 - causal status','Supplementary Table 8 | Genetic causal-status screen underlying the graded causal-status determinant',
  'Per-feature cis-Mendelian-randomization causal-status screen. 153 rows = 9 transcripts + 144 protein-outcome rows spanning 123 unique proteins '
  '(107 causal for a single non-eGFR cardiovascular-kidney-metabolic outcome + 16 for two or more). Transcript gate: Bonferroni-significant, Steiger-forward, F>=10 (GTEx-liver eQTL); '
  'protein gate: FDR q<0.10, non-MHC (UKB-PPP/Olink pQTL); eGFR-only nominations excluded a priori. MR beta and q are from the companion cis-MR screen; coloc PP.H4 and PP.H4/PP.H3 are '
  'populated only for coloc-staged rows (remainder MR-only, shown as "NA (not coloc-staged)"). Source: supplementary_tables/causal_status_supp_table.tsv.',
  s8,[13,13,9,16,15,9,9,10,10,11,12,46,14,13,15,13],
  center_cols={'Layer','Outcome','n instr','F-stat','MR β','MR P','MR q','Coloc PP.H4','PP.H4/PP.H3','# non-eGFR causal outcomes','In determinant','Panels measured','Coloc-staged in the companion cis-MR screen (ref. 17)'})

# ---------- S9: equivalence testing (TOST) (B5; 13 rows) ----------
to=RD('determinant_tost.tsv').copy()
to['Feature']=to['feature'].map(NAME)
to['Pooled β (SD)']=to['pooled_beta'].round(4); to['SE']=to['se'].round(4)
to['90% CI']=to.apply(lambda r:f"({r.ci90_lo:+.4f}, {r.ci90_hi:+.4f})",axis=1)
to['TOST P (±0.05 SD)']=to['tost_p_05'].apply(_sci); to['Equivalent ±0.05']=to['equiv_05'].map(_yn)
to['TOST P (±0.03 SD)']=to['tost_p_03'].apply(_sci); to['Equivalent ±0.03']=to['equiv_03'].map(_yn)
to['I² (%)']=to['I2'].round(1)
# Genetic causal status is a rare 0/1 binary (9 transcripts): its TOST passes only because the
# standardized coefficient is compressed by the predictor's tiny SD. The subtitle already declares that
# result non-interpretable, so a verdict cell reading "Yes" contradicts the sheet's own prose (and the
# manuscript). Mark it NA and DERIVE the continuous-feature counts so they cannot drift from the cells.
_rb=to['feature'].eq('causal_nonEGFR')
assert _rb.sum()==1, "expected exactly one rare-binary (causal_nonEGFR) row in the TOST table"
to.loc[_rb,'Equivalent ±0.05']='NA — rare binary'; to.loc[_rb,'Equivalent ±0.03']='NA — rare binary'
_cont=to[~_rb]; _k=len(_cont)
_n05=int((_cont['Equivalent ±0.05']=='Yes').sum()); _n03=int((_cont['Equivalent ±0.03']=='Yes').sum())
_incon=', '.join(_cont.loc[_cont['Equivalent ±0.05']!='Yes','Feature'])
assert (_n05,_n03,_k)==(11,9,12), f"TOST counts moved: {_n05}/{_k} at ±0.05, {_n03}/{_k} at ±0.03"
assert _incon=='Tissue-specificity (τ)', f"inconclusive continuous feature changed: {_incon!r}"
s9=to[['Feature','Pooled β (SD)','SE','90% CI','TOST P (±0.05 SD)','Equivalent ±0.05',
       'TOST P (±0.03 SD)','Equivalent ±0.03','I² (%)']]
write_sheet(wb,'S9 - equivalence (TOST)','Supplementary Table 9 | Formal practical-equivalence testing (two one-sided tests, TOST)',
  'Equivalence of each transcriptome determinant coefficient to zero against a symmetric SESOI of ±0.05 SD (primary) and ±0.03 SD (secondary), specified for the equivalence analysis. '
  f'Across the continuous determinant coefficients, {_n05} of {_k} are statistically equivalent within ±0.05 SD (TOST P<0.05) and {_n03} of {_k} within ±0.03 SD; only {_incon.lower()} is inconclusive '
  '(I²=96.9%, CI too wide), consistent with its dynamic-range-covariate role. Genetic causal status is displayed for completeness but is a rare 0/1 binary (9 transcripts) interpreted separately '
  '(Supplementary Table 14; Fig S8 / Additional file 10): its 90% CI lies inside the SESOI only because the standardized coefficient is compressed by the small predictor variance, so its equivalence '
  'verdict is reported as NA rather than Yes; on the interpretable '
  'raw scale it is −0.18 SD (95% CI −0.37 to +0.01) and underpowered — not shown equivalent to zero. Source: supplementary_tables/determinant_tost.tsv.',
  s9,[30,14,11,22,17,18,17,18,9],
  center_cols={'Pooled β (SD)','SE','TOST P (±0.05 SD)','Equivalent ±0.05','TOST P (±0.03 SD)','Equivalent ±0.03','I² (%)'})

# ---------- S10: determinant null under 3 measurability modes (B7; 39 rows) ----------
mo=RD('determinant_modes.tsv').copy()
MODEN={'unadjusted':'Unadjusted','covariate':'Measurability covariate','residualized':'Residualized (primary)'}
mo['Feature']=mo['feature'].map(NAME); mo['Mode']=mo['mode'].map(MODEN)
mo['Pooled β (SD)']=mo['pooled_beta'].round(4)
mo['95% CI']=mo.apply(lambda r:f"({r.pooled_beta-r.ci_halfwidth:+.4f}, {r.pooled_beta+r.ci_halfwidth:+.4f})",axis=1)
mo['I² (%)']=mo['I2'].round(1)
mo['_mo']=mo['mode'].map({'unadjusted':0,'covariate':1,'residualized':2})
mo=mo.sort_values(['Feature','_mo'])
s10=mo[['Feature','Mode','k','Pooled β (SD)','95% CI','I² (%)']]
write_sheet(wb,'S10 - determinant modes','Supplementary Table 10 | Determinant null under three measurability-adjustment schemes',
  'Determinant meta-analysis re-estimated under three measurability-adjustment schemes: unadjusted, measurability as a covariate, and measurability-residualized outcome (the canonical primary). '
  'The genetic causal-status standardized coefficient is stable across all three modes (β = −0.003 / −0.004 / −0.004); this standardized value is compressed by the rare-binary predictor and is not the primary interpretable contrast for this rare binary predictor and is not evaluated against the continuous-feature equivalence bound (interpretable raw contrast −0.18 SD; Supplementary Table 14), and no continuous feature reaches a moderate effect under any scheme '
  '(global max |β| = 0.040, druggability unadjusted). The two annotation features exceeding 0.03 SD without residualization (druggability, secreted-protein status) are abundance-correlated and shrink '
  'once measurability is modelled - residualization removes an abundance confound rather than manufacturing the null. Source: supplementary_tables/determinant_modes.tsv.',
  s10,[30,24,7,14,22,9],
  center_cols={'Mode','k','Pooled β (SD)','I² (%)'})

# ---------- S11: sensitivity (complete-case + nominatable universe) + missingness (B8) ----------
cc=RD('determinant_completecase.tsv').copy(); ccm=cc[cc['scope']=='META_completecase'].copy()
ccm['Feature']=ccm['feature'].map(NAME)
ccm['Pooled β (SD)']=ccm['beta'].round(4)
ccm['95% CI']=ccm.apply(lambda r:f"({r.beta-r.ci_halfwidth:+.4f}, {r.beta+r.ci_halfwidth:+.4f})",axis=1)
ccm['Cochran Q']=ccm['Q'].round(2); ccm['I² (%)']=ccm['I2'].round(1)
ccm['n (median)']=ccm['n'].astype(int); ccm['n (min)']=ccm['n_min'].astype(int)
b1=ccm[['Feature','k','Pooled β (SD)','95% CI','Cochran Q','I² (%)','n (median)','n (min)']]
ws11=write_sheet(wb,'S11 - sensitivity + missingness','Supplementary Table 11 | Determinant sensitivity (complete-case, nominatable-universe) and intrinsic-feature missingness',
  'Robustness of the determinant null to imputation and to the causal-status contrast definition, plus the intrinsic-feature missingness floor. '
  '(A) Complete-case meta-analysis - genes with LOEUF and τ observed, no imputation (10,681-16,630 genes/panel): genetic causal-status standardized β = −0.004 (95% CI [−0.008, −0.001]) ≈ the primary standardized estimate (compressed rare-binary coefficient, not the interpretable contrast; interpretable raw contrast −0.18 SD, Supplementary Table 14), '
  'so the null is not an imputation artifact; τ +0.030 (I²≈97%) is the known sign-blind dynamic-range effect, not a determinant. '
  '(B) Nominatable-universe contrast - restricted to the 1,432 cis-instrumented MR-tested genes (nominated vs tested-but-not-nominated; not-testable excluded): causal-status standardized β = −0.014 multivariable / −0.009 univariable (interpretable raw contrast −0.12 SD, 95% CI −0.32 to +0.08, Supplementary Table 14), '
  'CI includes zero from power loss (~8 nominated genes/panel) = a point-estimate consistency check, not a precisely-estimated null. '
  '(C) Missingness floor: architecture and causal-status annotations are available for only ~5-6% of genes/panel (missing 93.5-95.0%); LOEUF missing 13.3-43.5% (25.2% pooled); τ 2.9-26.5% (14.5% pooled). '
  'Sources: supplementary_tables/determinant_completecase.tsv, causal_nominatable_universe.tsv, intrinsic_missingness.tsv.',
  b1,[30,7,14,22,12,10,12,12],center_cols={'k','Pooled β (SD)','Cochran Q','I² (%)','n (median)','n (min)'})
ws11.cell(3,1,'(A) Complete-case meta-analysis (genes with LOEUF & τ observed; no imputation)').font=NOTE_FONT
# --- Block B: nominatable universe (2 META rows) ---
nu=RD('causal_nominatable_universe.tsv').copy()
rmv=nu[nu['scope']=='META_nominatable_mv'].iloc[0]; run=nu[nu['scope']=='META_nominatable_uni'].iloc[0]
b2=pd.DataFrame([
  {'Feature':'Genetic causal status','Contrast':'Multivariable','Pooled β (SD)':round(rmv['beta_mv'],4),
   '95% CI':f"({rmv['beta_mv']-rmv['ci_halfwidth']:+.4f}, {rmv['beta_mv']+rmv['ci_halfwidth']:+.4f})",
   'k':int(rmv['k']),'n universe':int(rmv['n_universe']),'n nominated (median)':int(rmv['n_nominated'])},
  {'Feature':'Genetic causal status','Contrast':'Univariable','Pooled β (SD)':round(run['beta_uni'],4),
   '95% CI':f"({run['beta_uni']-run['ci_halfwidth']:+.4f}, {run['beta_uni']+run['ci_halfwidth']:+.4f})",
   'k':int(run['k']),'n universe':int(run['n_universe']),'n nominated (median)':int(run['n_nominated'])}])
r0=4+len(b1)+2
ws11.cell(r0,1,'(B) Nominatable-universe contrast (cis-instrumented MR-tested genes; nominated vs tested-but-not-nominated)').font=NOTE_FONT
for j,col in enumerate(b2.columns,1):
    c=ws11.cell(r0+1,j,col); c.fill=HDR_FILL; c.font=HDR_FONT; c.alignment=CENTER; c.border=BORDER
for i,(_,row) in enumerate(b2.iterrows(),r0+2):
    for j,col in enumerate(b2.columns,1):
        c=ws11.cell(i,j,row[col]); c.font=BODY
        c.alignment=CENTER if col not in ('Feature','Contrast') else LEFT
# --- Block C: intrinsic missingness (per-panel + pooled), pct-missing matrix ---
im=RD('intrinsic_missingness.tsv').copy()
im=im[im['scope']!='UNIQUE_UNION'].copy()  # per B8: use per-panel/pooled, not the rare-gene-inflated union
FEATORD=['loeuf','tau','architecture','causal_status','is_secreted','is_enzyme','is_membrane']
mm=im.pivot(index='scope',columns='feature',values='pct_missing')[FEATORD]
panelord=[p for p in ['liver_table','adipose_table','gse141221_adipose_table','gse273902_blood_table',
  'gse157585_muscle_table','gse43471_adipose_table','gse83352_muscle_exercise_table',
  'gse106737_liver_bariatric_table','gse77962_adipose_diet_table','POOLED_SUM'] if p in mm.index]
mm=mm.reindex(panelord).reset_index()
mm.columns=['Panel / scope','LOEUF','τ','Architecture','Causal status','Secreted','Enzyme','Membrane']
r1=r0+2+len(b2)+2
ws11.cell(r1,1,'(C) Intrinsic-feature missingness by panel (% of genes missing each feature)').font=NOTE_FONT
for j,col in enumerate(mm.columns,1):
    c=ws11.cell(r1+1,j,col); c.fill=HDR_FILL; c.font=HDR_FONT; c.alignment=CENTER; c.border=BORDER
for i,(_,row) in enumerate(mm.iterrows(),r1+2):
    for j,col in enumerate(mm.columns,1):
        c=ws11.cell(i,j,row[col]); c.font=BODY
        c.alignment=LEFT if col=='Panel / scope' else CENTER
ws11.auto_filter.ref=None  # multi-block sheet: one range would filter across block boundaries
ws11.cell(1,3,'No AutoFilter: this sheet stacks several blocks, and a single filter range would '
               'span them.').font=NOTE_FONT

# ---------- S12: variance partition (B4; context reframe) ----------
vp=RD('variance_partition.tsv').copy()
TERMN={'same_tissue':'Shared tissue','same_intervention_family':'Shared intervention family','same_platform':'Shared assay platform'}
vt=vp[vp['term'].isin(TERMN)].copy(); vt['ord']=vt['term'].map({'same_tissue':0,'same_intervention_family':1,'same_platform':2})
vt=vt.sort_values('ord')
vt['Predictor']=vt['term'].map(TERMN)
vt['Unique R²']=vt['unique_R2'].round(4)
vt['Type-II (% total var)']=(vt['typeII_pct_totalvar']*100).round(2)
vt['Marginal R²']=vt['marginal_R2'].round(4)
vt['Jackknife 95% [min, max]']=vt.apply(lambda r:f"[{r.unique_R2_jack_min:.4f}, {r.unique_R2_jack_max:.4f}]",axis=1)
vt['Permutation P (unique)']=vt['perm_p_unique'].apply(_sci)
vt['n same pairs']=vt['n_same'].astype(int); vt['n diff pairs']=vt['n_diff'].astype(int)
s12=vt[['Predictor','Unique R²','Type-II (% total var)','Jackknife 95% [min, max]','Permutation P (unique)','Marginal R²','n same pairs','n diff pairs']]
# The legend's numbers are DERIVED from the same frame that fills the cells, never retyped: a
# hardcoded legend silently outlives a re-run of the engine (this is how "unique R²=0.005, P=0.27"
# survived the 2026-07-17 platform-label correction, and how 51.47->51.5->52 happened in v10).
_vpi=vp.set_index('term')
def _vp(term,col): return float(_vpi.loc[term,col])
def _supe(p):
    if p>=0.01: return f"{p:.2f}"
    e=math.floor(math.log10(p)); m=p/10**e
    ms=f"{round(m):d}" if abs(m-round(m))<0.05 else f"{m:.1f}"
    return ms+"×10⁻"+str(abs(e)).translate(str.maketrans('0123456789','⁰¹²³⁴⁵⁶⁷⁸⁹'))
_pl_r2, _pl_p = _vp('same_platform','unique_R2'), _vp('same_platform','perm_p_unique')
assert _pl_p>0.05, f"S12 legend calls the platform term a null, but perm P={_pl_p:.4g}"
ws12=write_sheet(wb,'S12 - variance partition','Supplementary Table 12 | Variance partition of pairwise reversal-signature correlations (context, not platform)',
  f"Partition of the variance in all 153 pairwise reversal-signature correlations (full-model R²={_vp('MODEL_full_R2','marginal_R2'):.2f}). Reversibility is organized by shared biological context: shared tissue and shared intervention family "
  f"each explain comparable, mutually independent variance (unique R²={_vp('same_tissue','unique_R2'):.2f} and {_vp('same_intervention_family','unique_R2'):.2f}; permutation P={_supe(_vp('same_tissue','perm_p_unique'))} and {_supe(_vp('same_intervention_family','perm_p_unique'))}), whereas shared assay platform explains essentially none (unique R²={_pl_r2:.4f}, P={_pl_p:.2f}, NS) - "
  'ruling out a batch/assay artifact. Context is co-organized by tissue and by intervention family, not by tissue alone and not by platform. Source: supplementary_tables/variance_partition.tsv.',
  s12,[26,12,20,26,22,13,13,13],
  center_cols={'Unique R²','Type-II (% total var)','Permutation P (unique)','Marginal R²','n same pairs','n diff pairs'})
# cell-means block (2x2 super-additivity) + model R2
cellm=[('Neither shared (diff tissue, diff family)','cell_diffT_diffF_mean',90),
       ('Same intervention family only','cell_diffT_sameF_mean',23),
       ('Same tissue only','cell_sameT_diffF_mean',27),
       ('Both shared (same tissue & family)','cell_sameT_sameF_mean',13)]
vmap=dict(zip(vp['term'],vp['beta_ols']))
cdf=pd.DataFrame([{'Cell (2×2 context)':lbl,'Mean Spearman ρ':round(vmap[k],3),'n pairs':n} for lbl,k,n in cellm])
r0=4+len(s12)+2
ws12.cell(r0,1,'Context cell means (super-additive when tissue and intervention family are both shared):').font=NOTE_FONT
for j,col in enumerate(cdf.columns,1):
    c=ws12.cell(r0+1,j,col); c.fill=HDR_FILL; c.font=HDR_FONT; c.alignment=CENTER; c.border=BORDER
for i,(_,row) in enumerate(cdf.iterrows(),r0+2):
    for j,col in enumerate(cdf.columns,1):
        c=ws12.cell(i,j,row[col]); c.font=BODY; c.alignment=LEFT if col=='Cell (2×2 context)' else CENTER
mr2=float(vp[vp['term']=='MODEL_full_R2']['marginal_R2'].iloc[0])
ws12.cell(r0+2+len(cdf)+1,1,f"Full-model R² = {mr2:.3f} (shared_commonality R² = {float(vp[vp['term']=='shared_commonality_R2']['unique_R2'].iloc[0]):.3f}).").font=NOTE_FONT
ws12.auto_filter.ref=None  # multi-block sheet, as S11
ws12.cell(1,3,'No AutoFilter: this sheet stacks several blocks, and a single filter range would '
               'span them.').font=NOTE_FONT

# ---------- S13: downsampling / detectability sensitivity (B6; softened wording) ----------
ds=RD('downsampling_sensitivity.tsv').copy()
METN={'raw_absbeta_vs_score':'Raw |β| vs score (direction-agnostic)','raw_signed_vs_score':'Raw signed β vs score (direction-aware)',
      'downsample_effN':'Downsampled effective-n re-rank','anchor_blom_vs_rawabs':'Anchor: blom(β) vs raw |β|'}
ds['Metric']=ds['metric'].map(METN)
ds['Spearman ρ']=ds['rho'].round(3)
ds['Downsample fraction']=ds['frac'].apply(lambda x:'' if pd.isna(x) else f"{x:g}")
ds['n replicates']=ds['n_replicates'].apply(lambda x:'' if pd.isna(x) else int(x))
s13=ds[['Metric','panel','Spearman ρ','n','Downsample fraction','n replicates','note']].copy()
s13.columns=['Metric','Panel / scope','Spearman ρ','n marks','Downsample fraction','n replicates','Note']
write_sheet(wb,'S13 - downsampling','Supplementary Table 13 | Downsampling / detectability sensitivity of the responsiveness ranking',
  'The responsiveness ranking is predominantly driven by the real signed reversal effect rather than dominated by detectability. A direction-aware raw-effect ranking reproduces the score '
  '(Spearman ρ=0.84 across 90,666 marks; 0.80 for multi-panel marks); the primary anti-artifact evidence is that the score itself correlates negatively with statistical power (corr(Z, log Σn)=−0.28), '
  'the opposite of a detectability artefact. Under simulated power reduction (standard-error-scaled Gaussian noise with model-based effective n, not subject-level resampling) the ranking degrades gracefully '
  '(ρ=0.83 at half effective n; no cliff), and a direction-agnostic magnitude ranking correlates only moderately (ρ=0.51, k≥2), consistent with the score rewarding cross-context direction-consistency. '
  'Source: supplementary_tables/downsampling_sensitivity.tsv.',
  s13,[40,30,12,11,18,13,60],
  center_cols={'Spearman ρ','n marks','Downsample fraction','n replicates'})

# ---------- S14: causal-status effect on reversibility (interpretable effect sizes) ----------
ce=RD('causal_status_effect.tsv').copy()
CENAME={'raw_group_smd_unadjusted':'Raw group SMD (unadjusted)',
 'multivariable_raw_coef':'Multivariable-adjusted raw coefficient',
 'nominatable_universe_smd':'Nominatable-universe SMD',
 'standardized_coef_anchor':'Standardized coefficient (canonical anchor)'}
ce['Estimate']=ce['estimate_name'].map(CENAME).fillna(ce['estimate_name'])
ce['Pooled']=ce['pooled'].round(4)
ce['95% CI']=ce.apply(lambda r:f"({r.ci_lo:+.4f}, {r.ci_hi:+.4f})",axis=1)
ce['k transcripts']=ce['k'].astype(int)
s14=ce[['Estimate','scale','Pooled','95% CI','k transcripts','note']].copy()
s14.columns=['Estimate','Scale','Pooled','95% CI','k','Note']
write_sheet(wb,'S14 - causal status effects','Supplementary Table 14 | Genetic causal-status effect on reversibility — interpretable effect sizes',
  'Effect of genetic causal status (a rare 0/1 predictor: 9 of ~18,600 transcripts) on transcriptome reversibility, reported on interpretable scales. '
  'These are the interpretable raw effect sizes underlying the reframed causal-status determinant. The raw group SMD is −0.18 SD (95% CI −0.37 to +0.01, includes zero); directionally concordant '
  'after multivariable adjustment (−0.19 SD) and in the fairer nominatable universe (−0.12 SD). The standardized anchor coefficient (−0.0036 SD) reproduces the canonical determinant meta-analysis but is '
  'compressed by the predictor SD (SD_x≈0.020) and is not the primary interpretable contrast for this rare binary predictor. The adjusted raw coefficient excluded zero nominally (−0.188, 95% CI −0.3534 to −0.0226), but that estimate rests on nine nominated genes, changed under universe restriction, and was not supported by the unadjusted or nominatable-universe contrasts; we therefore treat it as non-robust and underpowered rather than as evidence of a stable causal-status effect. No specification gives an association that is both significant and adequately powered: the estimates rest on nine genes with wide intervals, so genetic '
  'causal status is underpowered for reversibility rather than shown equivalent to zero (unlike the continuous intrinsic features). Source: supplementary_tables/causal_status_effect.tsv.',
  s14,[42,26,11,22,7,58],center_cols={'Pooled','k'})

# ---------- S15: true-weight-loss-only sensitivity (determinant null + tissue organization) ----------
wl=RD('sensitivity_wlonly/determinant_meta_wlonly.tsv').copy()
SCENN={'k9_canonical':'All 9 panels (anchor)','k6_strict':'Diet + surgery, strict (k=6; 3 tissues)',
       'k7_keepdietex':'Diet + surgery + diet-exercise (k=7)'}
wl['Scenario']=wl['scenario'].map(SCENN)
wl['Feature']=wl['feature'].map(NAME)
wl['Pooled β (SD)']=wl['pooled_beta'].round(4)
wl['95% CI']=wl.apply(lambda r:f"({r.ci_lo:+.4f}, {r.ci_hi:+.4f})",axis=1)
wl['I² (%)']=wl['I2'].round(1)
wl['Reaches ±0.05 SESOI']=wl['exceeds_SESOI'].map(_yn)
wl['_o']=wl['scenario'].map({'k9_canonical':0,'k6_strict':1,'k7_keepdietex':2})
wl=wl.sort_values(['_o','Feature'])
s15=wl[['Scenario','k','Feature','Pooled β (SD)','95% CI','I² (%)','Reaches ±0.05 SESOI']]
ws15=write_sheet(wb,'S15 - weight-loss-only','Supplementary Table 15 | True-weight-loss-only sensitivity (determinant null and tissue organization)',
  'The transcriptome determinant meta-analysis and the tissue/intervention variance partition recomputed after excluding every metformin, SGLT2-inhibitor, GLP-1-agonist and exercise panel, retaining only '
  'caloric-restriction, dietary and bariatric-surgery panels. Determinant block (below): per-feature Hartung-Knapp pooled coefficients for the full nine-panel meta (anchor; reproduces the canonical result to '
  '~1e-16), the strict six-panel diet-and-surgery set (three tissues: adipose/liver/blood) and a seven-panel variant retaining the single diet-plus-exercise panel. No feature reaches the ±0.05 SD SESOI on the '
  'point estimate in any set (largest |pooled β| = 0.029 SD, 1.7-fold below the bound); tissue-specificity (τ) and, in the strict set, the LOEUF-missing indicator remain heterogeneous (I²≈91–97%) and CI-wide, the '
  'known dynamic-range/annotation features rather than novel determinants. Genetic causal status stays small and its CI crosses zero in the strict set (compressed rare-binary standardized coefficient, not the interpretable contrast '
  'size; interpretable raw contrast −0.18 SD, 95% CI −0.37 to +0.01, nine transcripts; Supplementary Table 14). Empagliflozin (EMPEROR) is a proteome panel and never enters the transcriptome determinant meta. '
  'Context block (further below): within- versus cross-tissue mean reversal correlation and the shared-tissue/intervention/platform variance partition on the weight-loss-only panels, under a twelve-panel definition '
  '(diet+exercise retained as diet) and a strict eleven-panel definition (diet+exercise also dropped). Sources: supplementary_tables/determinant_meta_wlonly.tsv, supplementary_tables/context_wlonly.tsv.',
  s15,[34,7,30,14,22,9,20],center_cols={'k','Pooled β (SD)','I² (%)','Reaches ±0.05 SESOI'})
# --- context block ---
cw=RD('sensitivity_wlonly/context_wlonly.tsv').copy()
def _cv(scen,term,col):
    r=cw[(cw['scenario']==scen)&(cw['term']==term)]
    return r[col].iloc[0] if len(r) else float('nan')
CSCEN=[('lenient12','12 panels (diet+exercise retained)'),('strict11','11 panels (diet+exercise dropped)')]
cctx=pd.DataFrame([{
    'Panel definition':lbl,
    'Within-tissue mean ρ':f"{_cv(sc,'within_tissue_mean_rho','beta_ols'):+.3f} (n={int(_cv(sc,'within_tissue_mean_rho','n_same'))})",
    'Cross-tissue mean ρ':f"{_cv(sc,'cross_tissue_mean_rho','beta_ols'):+.3f} (n={int(_cv(sc,'cross_tissue_mean_rho','n_same'))})",
    'Shared-tissue unique R²':round(_cv(sc,'same_tissue','unique_R2'),3),
    'Shared-tissue perm P':_sci(_cv(sc,'same_tissue','perm_p_marginal_tissue')),
    'Shared-platform unique R²':round(_cv(sc,'same_platform','unique_R2'),4),
    'Shared-intervention-family unique R²':round(_cv(sc,'same_intervention_family','unique_R2'),3),
} for sc,lbl in CSCEN])
r0=4+len(s15)+2
ws15.cell(r0,1,'Context (tissue/intervention organization + variance partition) on the weight-loss-only panels:').font=NOTE_FONT
for j,col in enumerate(cctx.columns,1):
    c=ws15.cell(r0+1,j,col); c.fill=HDR_FILL; c.font=HDR_FONT; c.alignment=CENTER; c.border=BORDER
for i,(_,row) in enumerate(cctx.iterrows(),r0+2):
    for j,col in enumerate(cctx.columns,1):
        c=ws15.cell(i,j,row[col]); c.font=BODY; c.alignment=LEFT if col=='Panel definition' else CENTER
ws15.auto_filter.ref=None  # multi-block sheet, as S11
ws15.cell(1,3,'No AutoFilter: this sheet stacks several blocks, and a single filter range would '
               'span them.').font=NOTE_FONT

# ---------- Supp: restoration uncertainty (round-2: CIs / binomial-permutation P for the 52-54% contexts) ----------
ru=RD('restoration_uncertainty/restoration_uncertainty.tsv').copy()
ru=ru.sort_values('pct_toward_lean',ascending=False)
s16=pd.DataFrame({
    'Context':ru['context'], 'Tissue':ru['tissue'], 'n marks':ru['n_marks'],
    '% toward lean':ru['pct_toward_lean'].round(2),
    '95% CI (Wilson)':ru.apply(lambda r:f"({r.wilson_ci_lo:.1f}, {r.wilson_ci_hi:.1f})",axis=1),
    # Round-3: "very small P values print as 0.00e+00 … replace with '<1e-300' or the smallest
    # representable threshold used by the script." The ten zeros are NOT one quantity, and a blanket
    # "<1e-300" would be FALSE for nine of them:
    #   * binom_p_vs_50 / centred_binom_p_vs_50 are ANALYTIC exact-binomial P values -> a 0 is genuine
    #     double underflow (smallest nonzero in this column is 2.0e-59), so "<1e-300" is honest.
    #   * perm_p_vs_p0 is an EMPIRICAL permutation P over B = 10,000 draws, computed as a bare mean
    #     with no +1 correction (restoration_uncertainty.py L150), so it is BOUNDED BELOW BY 1/B.
    #     A 0 means "no draw was as extreme", i.e. P < 1e-4 — reporting it as <1e-300 would assert a
    #     precision the permutation cannot deliver.
    'P vs 50% (exact binomial)':ru['binom_p_vs_50'].map(_sci_analytic),
    'Distinguishable from 50%':ru['sig_vs_50'].map(_yn),
    'Marginal-conditioned baseline (%)':ru['chance_baseline_p0'].round(2),
    'P vs that baseline (permutation)':ru['perm_p_vs_p0'].map(_sci_perm),
    'Same verdict under both baselines':ru['agrees_across_baselines'].map(_yn),
    '% of reversal effects positive':ru['pct_rev_up'],
    'Verdict changes if the panel is median-centred':ru['flips_on_centring'].map(_yn),
})
write_sheet(wb,'S16 - restoration uncertainty',
  'Supplementary Table 16 | Uncertainty of the restoration (% toward lean) statistic, and why an interval on it is weak',
  'Confidence intervals, exact-binomial and permutation P values, and explicit near-random wording for the 52-54% contexts. '
  'Exact binomial vs the 50% line (the manuscript\'s baseline): four contexts are not distinguishable from chance — MS bariatric (55%, P = 0.12), metformin (54%, P = 0.23), '
  'empagliflozin/EMPEROR (52%, P = 0.31) and GSE273902 blood (51.5%, P = 0.10); all four are small-n plasma or blood panels, and all are marked with open circles in Fig. 4b. '
  'A SECOND baseline is reported for robustness only: panels differ in the marginal sign balance of their reversal vector (90% of EMPEROR plasma proteins rise — the documented '
  'plasma-volume-contraction effect of SGLT2 inhibition — versus 84% falling for metformin and 50% for semaglutide), so a random pairing against a sign-imbalanced BMI axis does not '
  'yield 50% but p0 = P(bmi>0)P(rev<0) + P(bmi<0)P(rev>0), here 47.0-53.2% (closed form; confirmed against a 10,000-fold gene-label permutation to 0.02 pp). '
  'p0 is a DIFFERENT ESTIMAND, not a correction to 50%: P(rev>0) is post-treatment and not ancillary, so conditioning on it conditions away part of a genuine global restorative shift. '
  'Only verdicts agreeing under both baselines are relied on; the two that disagree (metformin, blood) are marked and no claim rests on them. Empagliflozin is not distinguishable from '
  'chance under the 50% line (P = 0.31), under p0 (P = 0.20), or after median-centring (P = 0.09) — the one baseline-independent conclusion among the near-chance contexts. '
  'CAVEAT ON THE INSTRUMENT ITSELF: a binomial interval treats marks as independent replicates, which they are not, and the relevant replication unit is the subject, not the mark; '
  'a gene-label permutation cannot repair this because it randomises gene identity, so inter-mark correlation cannot enter its null by construction '
  '(its SD is deterministic given the margins, perm_SD/binom_SD = 4*sqrt(a(1-a)b(1-b))). A correlation-respecting interval would require subject-level resampling, which these '
  'published per-mark summary panels do not permit. The percentages therefore anchor to the primary analysis scripts (max |delta| 0.43 pp) but are reported with "near-random" wording '
  'rather than as precisely-bounded estimates. Source: supplementary_tables/restoration_uncertainty.tsv.',
  s16,[26,10,9,13,16,20,16,20,20,20,16,20],
  center_cols=set(s16.columns)-{'Context'})

# ---------- Contents (round 6, 9.3: "add a Contents / data-dictionary sheet") ----------
# DERIVED from the workbook that was just built: the sheet list and every title are read back out
# of wb, so a Contents row cannot describe a sheet that is not there, and a new sheet cannot be
# omitted. Placed first.
_contents=pd.DataFrame(
    [{'Sheet':s, 'Table': (wb[s]['A1'].value or '').split('|')[0].strip(),
      'What it contains': (wb[s]['A1'].value or '').split('|',1)[-1].strip(),
      'Rows': max(0, wb[s].max_row-4)}
     for s in wb.sheetnames])
assert len(_contents)==len(wb.sheetnames) and _contents['Table'].ne('').all(), 'Contents derivation failed'
_ws_c=write_sheet(wb,'Contents','Contents | Paper B tables workbook',
  "Every sheet in this workbook, in order. Main Tables 1 and 2 are also supplied as Word tables in the manuscript; this workbook is the machine-readable form and carries the full per-panel registry "
  "and Supplementary Tables S1-S16. Each sheet carries its own title in row 1 and a note in row 2 giving the method, the interpretation rule where one applies, and the results file it was generated from. "
  "Rows counts the data rows below each sheet's header. Sheets that stack several blocks (S11, S12, S15) carry no AutoFilter, because a single filter range would span the blocks.",
  _contents,[30,26,86,8], center_cols={'Rows'})
wb.move_sheet('Contents', offset=-(len(wb.sheetnames)-1))
assert wb.sheetnames[0]=='Contents', wb.sheetnames[:3]
# Assert AFTER the move as well: Contents must describe every other sheet, so adding a sheet later
# without regenerating cannot leave a silently incomplete index.
_listed=set(_contents['Sheet']); _actual=set(wb.sheetnames)-{'Contents'}
assert _listed==_actual, f"Contents misses {sorted(_actual-_listed)}; lists absent {sorted(_listed-_actual)}"

wb.properties.creator='Bertrand Chin-Ming Tan'; wb.properties.lastModifiedBy='Bertrand Chin-Ming Tan'
wb.save(os.path.join(HERE,'Paper_B_Tables.xlsx'))
print('wrote manuscript/tables/Paper_B_Tables.xlsx with sheets:', wb.sheetnames)
print('ED1 rows=%d | Table1(det) rows=%d | core=%d | drugclass=%d'%(len(ed1),len(t2),len(core),len(dc)))
print('S8=%d S9=%d S10=%d S11:complete-case=%d nominatable=%d missingness=%d S12=%d S13=%d S14=%d S15=%d'%(
    len(s8),len(s9),len(s10),len(b1),len(b2),len(mm),len(s12),len(s13),len(s14),len(s15)))
